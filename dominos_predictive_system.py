"""
============================================================
Domino's Predictive Purchase Order System
============================================================
Capstone Project — Food Service Industry
Author  : Data Science Pipeline
Version : 1.0  (production-ready)

Pipeline
--------
1.  Data Ingestion & Validation
2.  Data Cleaning & Preprocessing
3.  Exploratory Data Analysis  → saves plots
4.  Feature Engineering
5.  Time-Series Forecasting    → Prophet per pizza_name_id
6.  Ingredient Calculation
7.  Purchase Order Generation  → Excel + CSV

Usage
-----
    python dominos_predictive_system.py

Outputs (all in ./outputs/)
-------
    eda_*.png              – EDA visualisations
    model_evaluation.csv   – per-SKU MAPE
    weekly_forecast.csv    – 7-day quantity forecast
    purchase_order.xlsx    – formatted purchase order workbook
    purchase_order.csv     – plain-text purchase order
"""

# ── Standard library ──────────────────────────────────────────────────────────
import os
import sys
import warnings
import logging
from pathlib import Path
from datetime import datetime, timedelta

# ── Third-party ───────────────────────────────────────────────────────────────
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")               # non-interactive backend (safe for any env)
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
from prophet import Prophet
from sklearn.metrics import mean_absolute_percentage_error

warnings.filterwarnings("ignore")
logging.getLogger("prophet").setLevel(logging.ERROR)
logging.getLogger("cmdstanpy").setLevel(logging.ERROR)

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
DATA_DIR    = BASE_DIR
OUTPUT_DIR  = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

SALES_FILE       = DATA_DIR / "Pizza_Sale.xlsx"
INGREDIENT_FILE  = DATA_DIR / "Pizza_ingredients.xlsx"

# ── Config ────────────────────────────────────────────────────────────────────
FORECAST_DAYS   = 7          # next-week horizon
SAFETY_BUFFER   = 0.10       # 10 % buffer on ingredient quantities
MIN_HISTORY_OBS = 20         # minimum daily obs per SKU to train Prophet
RANDOM_SEED     = 42

# ── Colour palette (consistent across all plots) ──────────────────────────────
PALETTE = {
    "Classic" : "#E63946",
    "Supreme" : "#457B9D",
    "Veggie"  : "#2A9D8F",
    "Chicken" : "#F4A261",
    "Unknown" : "#9B9B9B",
}

np.random.seed(RANDOM_SEED)

# =============================================================================
# 0.  LOGGING
# =============================================================================
logging.basicConfig(
    level   = logging.INFO,
    format  = "%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt = "%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(OUTPUT_DIR / "run.log", mode="w"),
    ],
)
log = logging.getLogger(__name__)


# =============================================================================
# 1.  DATA INGESTION & VALIDATION
# =============================================================================
def load_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load & validate both Excel files."""
    log.info("Loading sales data …")
    sales = pd.read_excel(SALES_FILE)

    log.info("Loading ingredient data …")
    ingr  = pd.read_excel(INGREDIENT_FILE)

    log.info("Sales   : %d rows × %d cols", *sales.shape)
    log.info("Ingredients: %d rows × %d cols", *ingr.shape)
    return sales, ingr


# =============================================================================
# 2.  DATA CLEANING & PREPROCESSING
# =============================================================================
def clean_sales(df: pd.DataFrame) -> pd.DataFrame:
    """Clean and type-cast the sales dataframe."""
    log.info("Cleaning sales data …")
    df = df.copy()

    # ── Date parsing ──────────────────────────────────────────────────────────
    df["order_date"] = pd.to_datetime(df["order_date"], errors="coerce")
    null_dates = df["order_date"].isna().sum()
    if null_dates:
        log.warning("Dropped %d rows with unparseable dates", null_dates)
        df = df.dropna(subset=["order_date"])

    # ── Numeric coercion ──────────────────────────────────────────────────────
    for col in ["quantity", "unit_price", "total_price"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # ── Quantity: drop non-positive ───────────────────────────────────────────
    bad_qty = (df["quantity"].isna()) | (df["quantity"] <= 0)
    if bad_qty.sum():
        log.warning("Dropped %d rows with invalid quantity", bad_qty.sum())
        df = df[~bad_qty]

    # ── pizza_name_id: fill missing from pizza_name ───────────────────────────
    missing_id = df["pizza_name_id"].isna()
    if missing_id.sum():
        log.warning("%d rows missing pizza_name_id — using pizza_name as proxy",
                    missing_id.sum())
        df.loc[missing_id, "pizza_name_id"] = (
            df.loc[missing_id, "pizza_name"]
              .str.lower().str.replace(r"[^a-z0-9]", "_", regex=True)
        )

    # ── Standardise category column ───────────────────────────────────────────
    df["pizza_category"] = df["pizza_category"].fillna("Unknown").str.strip()

    # ── Size normalisation ────────────────────────────────────────────────────
    df["pizza_size"] = df["pizza_size"].str.upper().fillna("M")

    # ── Total price sanity check ──────────────────────────────────────────────
    expected = (df["quantity"] * df["unit_price"]).round(2)
    mismatch = (df["total_price"].round(2) - expected).abs() > 0.05
    if mismatch.sum():
        log.warning("Recalculating total_price for %d mismatched rows", mismatch.sum())
        df.loc[mismatch, "total_price"] = expected[mismatch]

    df = df.reset_index(drop=True)
    log.info("Clean sales dataset: %d rows", len(df))
    return df


def clean_ingredients(df: pd.DataFrame) -> pd.DataFrame:
    """Clean the ingredient dataframe."""
    log.info("Cleaning ingredient data …")
    df = df.copy()

    df.columns = df.columns.str.strip()
    df["Items_Qty_In_Grams"] = pd.to_numeric(df["Items_Qty_In_Grams"], errors="coerce")

    # Fill missing ingredient quantities with 0 (warn)
    missing_qty = df["Items_Qty_In_Grams"].isna().sum()
    if missing_qty:
        log.warning("Filling %d missing ingredient quantities with 0", missing_qty)
        df["Items_Qty_In_Grams"] = df["Items_Qty_In_Grams"].fillna(0.0)

    df = df.dropna(subset=["pizza_name_id", "pizza_ingredients"])
    df["pizza_name_id"]    = df["pizza_name_id"].str.strip()
    df["pizza_ingredients"] = df["pizza_ingredients"].str.strip()
    log.info("Clean ingredients: %d rows, %d unique SKUs",
             len(df), df["pizza_name_id"].nunique())
    return df


# =============================================================================
# 3.  EXPLORATORY DATA ANALYSIS
# =============================================================================
def run_eda(df: pd.DataFrame) -> None:
    """Generate and save all EDA plots."""
    log.info("Running EDA …")
    sns.set_style("whitegrid")
    plt.rcParams.update({"font.size": 11, "figure.dpi": 150})

    # 3-a  Daily revenue trend ─────────────────────────────────────────────────
    daily_rev = (df.groupby("order_date")["total_price"]
                   .sum()
                   .reset_index()
                   .rename(columns={"total_price": "revenue"}))

    fig, ax = plt.subplots(figsize=(14, 4))
    ax.plot(daily_rev["order_date"], daily_rev["revenue"],
            color="#E63946", linewidth=0.8, alpha=0.7)
    ax.fill_between(daily_rev["order_date"], daily_rev["revenue"],
                    alpha=0.15, color="#E63946")

    # 7-day rolling average
    daily_rev["rolling_7"] = daily_rev["revenue"].rolling(7).mean()
    ax.plot(daily_rev["order_date"], daily_rev["rolling_7"],
            color="#1D3557", linewidth=1.8, label="7-day MA")

    ax.set(title="Daily Revenue Trend (2015)", xlabel="Date", ylabel="Revenue ($)")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b"))
    ax.xaxis.set_major_locator(mdates.MonthLocator())
    ax.legend()
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "eda_01_daily_revenue_trend.png")
    plt.close()
    log.info("  Plot saved: eda_01_daily_revenue_trend.png")

    # 3-b  Sales by category ───────────────────────────────────────────────────
    cat_rev = (df.groupby("pizza_category")["total_price"]
                  .sum()
                  .sort_values(ascending=False))

    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    colours = [PALETTE.get(c, "#9B9B9B") for c in cat_rev.index]

    cat_rev.plot(kind="bar", ax=axes[0], color=colours, edgecolor="white")
    axes[0].set(title="Revenue by Category", xlabel="", ylabel="Revenue ($)")
    axes[0].tick_params(axis="x", rotation=30)
    for p in axes[0].patches:
        axes[0].annotate(f"${p.get_height()/1e3:.0f}K",
                         (p.get_x() + p.get_width()/2., p.get_height()),
                         ha="center", va="bottom", fontsize=9)

    # Pie chart
    axes[1].pie(cat_rev.values, labels=cat_rev.index,
                colors=colours, autopct="%1.1f%%", startangle=140,
                wedgeprops={"edgecolor":"white","linewidth":1.5})
    axes[1].set_title("Category Revenue Share")
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "eda_02_sales_by_category.png")
    plt.close()
    log.info("  Plot saved: eda_02_sales_by_category.png")

    # 3-c  Top-20 pizzas by quantity ───────────────────────────────────────────
    top20 = (df.groupby("pizza_name")["quantity"]
               .sum()
               .sort_values(ascending=False)
               .head(20))

    fig, ax = plt.subplots(figsize=(12, 7))
    top20.sort_values().plot(kind="barh", ax=ax, color="#457B9D", edgecolor="white")
    ax.set(title="Top 20 Pizzas by Units Sold", xlabel="Units Sold", ylabel="")
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:,.0f}"))
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "eda_03_top20_pizzas.png")
    plt.close()
    log.info("  Plot saved: eda_03_top20_pizzas.png")

    # 3-d  Monthly seasonality ─────────────────────────────────────────────────
    df["month"] = df["order_date"].dt.month
    monthly = (df.groupby(["month", "pizza_category"])["quantity"]
                 .sum()
                 .reset_index())

    fig, ax = plt.subplots(figsize=(12, 5))
    for cat, grp in monthly.groupby("pizza_category"):
        ax.plot(grp["month"], grp["quantity"],
                label=cat, color=PALETTE.get(cat, "#9B9B9B"),
                marker="o", linewidth=2)
    ax.set(title="Monthly Sales by Category",
           xlabel="Month", ylabel="Units Sold")
    ax.set_xticks(range(1, 13))
    ax.set_xticklabels(["Jan","Feb","Mar","Apr","May","Jun",
                         "Jul","Aug","Sep","Oct","Nov","Dec"])
    ax.legend()
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "eda_04_monthly_seasonality.png")
    plt.close()
    log.info("  Plot saved: eda_04_monthly_seasonality.png")

    # 3-e  Day-of-week pattern ────────────────────────────────────────────────
    df["dow"] = df["order_date"].dt.day_name()
    dow_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    dow_qty   = df.groupby("dow")["quantity"].sum().reindex(dow_order)

    fig, ax = plt.subplots(figsize=(9, 4))
    bars = ax.bar(dow_qty.index, dow_qty.values,
                  color=["#E63946" if d in ["Friday","Saturday","Sunday"]
                          else "#457B9D" for d in dow_qty.index],
                  edgecolor="white")
    ax.set(title="Sales by Day of Week", xlabel="", ylabel="Units Sold")
    ax.tick_params(axis="x", rotation=20)
    for bar, val in zip(bars, dow_qty.values):
        ax.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 50,
                f"{val:,.0f}", ha="center", va="bottom", fontsize=8)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "eda_05_dayofweek_pattern.png")
    plt.close()
    log.info("  Plot saved: eda_05_dayofweek_pattern.png")

    # 3-f  Order-hour heatmap ─────────────────────────────────────────────────
    try:
        df["hour"] = pd.to_datetime(df["order_time"], format="%H:%M:%S",
                                    errors="coerce").dt.hour
        pivot = (df.groupby(["dow","hour"])["quantity"]
                   .sum()
                   .unstack(fill_value=0)
                   .reindex(dow_order))

        fig, ax = plt.subplots(figsize=(14, 5))
        sns.heatmap(pivot, cmap="YlOrRd", linewidths=0.3,
                    linecolor="grey", ax=ax, fmt=".0f", annot=False)
        ax.set(title="Orders Heatmap: Day × Hour",
               xlabel="Hour of Day", ylabel="")
        plt.tight_layout()
        plt.savefig(OUTPUT_DIR / "eda_06_heatmap_day_hour.png")
        plt.close()
        log.info("  Plot saved: eda_06_heatmap_day_hour.png")
    except Exception as exc:
        log.warning("  Heatmap skipped: %s", exc)

    # 3-g  Price distribution by size ─────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(10, 4))
    size_order = ["S", "M", "L", "XL", "XXL"]
    sizes_present = [s for s in size_order if s in df["pizza_size"].unique()]
    for size in sizes_present:
        subset = df[df["pizza_size"] == size]["unit_price"].dropna()
        ax.hist(subset, bins=30, alpha=0.6, label=size)
    ax.set(title="Unit Price Distribution by Size",
           xlabel="Unit Price ($)", ylabel="Frequency")
    ax.legend()
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "eda_07_price_distribution.png")
    plt.close()
    log.info("  Plot saved: eda_07_price_distribution.png")

    log.info("EDA complete — all plots saved to %s", OUTPUT_DIR)


# =============================================================================
# 4.  FEATURE ENGINEERING
# =============================================================================
def build_daily_series(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate to daily SKU-level quantity series."""
    log.info("Building daily time series per SKU …")
    daily = (df.groupby(["order_date", "pizza_name_id"])["quantity"]
               .sum()
               .reset_index()
               .rename(columns={"order_date": "ds",
                                 "quantity"  : "y"}))

    # Add calendar features (for EDA reference — Prophet handles its own internally)
    daily["weekday"]  = daily["ds"].dt.dayofweek
    daily["month"]    = daily["ds"].dt.month
    daily["week"]     = daily["ds"].dt.isocalendar().week.astype(int)
    return daily


# =============================================================================
# 5.  TIME-SERIES FORECASTING  (Prophet)
# =============================================================================
def train_and_forecast(daily: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Train one Prophet model per SKU and forecast FORECAST_DAYS ahead.
    Returns (forecast_df, evaluation_df).
    """
    log.info("Training Prophet models …")

    skus        = daily["pizza_name_id"].unique()
    forecasts   = []
    eval_rows   = []

    # Determine forecast start (day after last known date)
    last_date      = daily["ds"].max()
    forecast_start = last_date + timedelta(days=1)
    forecast_end   = forecast_start + timedelta(days=FORECAST_DAYS - 1)

    for idx, sku in enumerate(skus, 1):
        sku_df = (daily[daily["pizza_name_id"] == sku][["ds", "y"]]
                  .sort_values("ds")
                  .reset_index(drop=True))

        if len(sku_df) < MIN_HISTORY_OBS:
            # Fall back to mean for very sparse SKUs
            mean_qty = sku_df["y"].mean()
            log.debug("  SKU %s — sparse (%d obs), using mean %.2f",
                      sku, len(sku_df), mean_qty)
            for d in pd.date_range(forecast_start, forecast_end):
                forecasts.append({
                    "pizza_name_id" : sku,
                    "ds"            : d,
                    "yhat"          : max(0, mean_qty),
                    "yhat_lower"    : max(0, mean_qty * 0.7),
                    "yhat_upper"    : mean_qty * 1.3,
                    "method"        : "mean_fallback",
                })
            eval_rows.append({"pizza_name_id": sku, "mape": np.nan,
                               "method": "mean_fallback"})
            continue

        # ── Walk-forward validation: hold out last 7 days ────────────────────
        cutoff     = sku_df["ds"].max() - timedelta(days=7)
        train_df   = sku_df[sku_df["ds"] <= cutoff]
        val_df     = sku_df[sku_df["ds"] >  cutoff]

        mape_val = np.nan
        if len(train_df) >= MIN_HISTORY_OBS and len(val_df) > 0:
            try:
                m_val = Prophet(
                    yearly_seasonality  = True,
                    weekly_seasonality  = True,
                    daily_seasonality   = False,
                    seasonality_mode    = "multiplicative",
                    changepoint_prior_scale = 0.1,
                    interval_width      = 0.90,
                )
                m_val.fit(train_df)
                future_val = m_val.make_future_dataframe(periods=7)
                fc_val     = m_val.predict(future_val)
                fc_val_7   = fc_val[fc_val["ds"].isin(val_df["ds"])]
                preds      = fc_val_7["yhat"].clip(lower=0).values
                actuals    = val_df["y"].values
                if actuals.sum() > 0:
                    mape_val = mean_absolute_percentage_error(actuals, preds)
            except Exception:
                pass

        # ── Full model on all data ────────────────────────────────────────────
        try:
            m = Prophet(
                yearly_seasonality      = True,
                weekly_seasonality      = True,
                daily_seasonality       = False,
                seasonality_mode        = "multiplicative",
                changepoint_prior_scale = 0.1,
                interval_width          = 0.90,
            )
            m.fit(sku_df)
            future = m.make_future_dataframe(periods=FORECAST_DAYS)
            fc     = m.predict(future)
            week_fc = fc[fc["ds"] > last_date].copy()
            week_fc["pizza_name_id"] = sku
            week_fc["yhat"]          = week_fc["yhat"].clip(lower=0)
            week_fc["yhat_lower"]    = week_fc["yhat_lower"].clip(lower=0)
            week_fc["method"]        = "prophet"

            for _, row in week_fc.iterrows():
                forecasts.append({
                    "pizza_name_id" : sku,
                    "ds"            : row["ds"],
                    "yhat"          : row["yhat"],
                    "yhat_lower"    : row["yhat_lower"],
                    "yhat_upper"    : row["yhat_upper"],
                    "method"        : "prophet",
                })

            eval_rows.append({"pizza_name_id": sku, "mape": mape_val,
                               "method": "prophet"})

        except Exception as exc:
            log.error("  SKU %s — Prophet failed: %s", sku, exc)
            mean_qty = sku_df["y"].mean()
            for d in pd.date_range(forecast_start, forecast_end):
                forecasts.append({
                    "pizza_name_id" : sku,
                    "ds"            : d,
                    "yhat"          : max(0, mean_qty),
                    "yhat_lower"    : max(0, mean_qty * 0.7),
                    "yhat_upper"    : mean_qty * 1.3,
                    "method"        : "mean_fallback",
                })
            eval_rows.append({"pizza_name_id": sku, "mape": np.nan,
                               "method": "mean_fallback"})

        if idx % 10 == 0:
            log.info("  Trained %d / %d SKUs …", idx, len(skus))

    forecast_df = pd.DataFrame(forecasts)
    eval_df     = pd.DataFrame(eval_rows)

    # ── Aggregate to total weekly quantities per SKU ───────────────────────────
    weekly_fc = (forecast_df.groupby("pizza_name_id")["yhat"]
                             .sum()
                             .reset_index()
                             .rename(columns={"yhat": "predicted_qty_week"}))
    weekly_fc["predicted_qty_week"] = weekly_fc["predicted_qty_week"].round().astype(int)
    weekly_fc["predicted_qty_week"] = weekly_fc["predicted_qty_week"].clip(lower=0)

    valid_mape = eval_df["mape"].dropna()
    overall_mape = valid_mape.mean()
    log.info("Forecasting complete — overall mean MAPE: %.2f%%",
             overall_mape * 100 if not np.isnan(overall_mape) else float("nan"))

    return weekly_fc, eval_df, forecast_df


def plot_forecast_sample(daily: pd.DataFrame,
                          forecast_full: pd.DataFrame,
                          n_samples: int = 6) -> None:
    """Plot forecast vs actuals for top-n SKUs."""
    top_skus = (daily.groupby("pizza_name_id")["y"]
                      .sum()
                      .nlargest(n_samples)
                      .index.tolist())

    fig, axes = plt.subplots(2, 3, figsize=(18, 8))
    axes = axes.flatten()

    for i, sku in enumerate(top_skus):
        ax    = axes[i]
        hist  = daily[daily["pizza_name_id"] == sku].sort_values("ds")
        fcast = forecast_full[forecast_full["pizza_name_id"] == sku].sort_values("ds")

        # Monthly aggregation for readability
        hist_m  = hist.set_index("ds")["y"].resample("W").sum().reset_index()
        fcast_m = fcast.set_index("ds")["yhat"].resample("W").sum().reset_index()

        ax.plot(hist_m["ds"],  hist_m["y"],     label="Actual",   color="#1D3557")
        ax.plot(fcast_m["ds"], fcast_m["yhat"], label="Forecast", color="#E63946",
                linestyle="--")
        ax.set_title(sku.replace("_", " ").title(), fontsize=9)
        ax.set_ylabel("Units / Week")
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b"))
        ax.xaxis.set_major_locator(mdates.MonthLocator(bymonth=[1,4,7,10,12]))
        ax.tick_params(axis="x", rotation=20, labelsize=7)
        if i == 0:
            ax.legend(fontsize=8)

    plt.suptitle("Prophet Forecast — Top SKUs (weekly aggregation)", y=1.01)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "eda_08_prophet_forecast_sample.png",
                bbox_inches="tight")
    plt.close()
    log.info("  Plot saved: eda_08_prophet_forecast_sample.png")


# =============================================================================
# 6.  INGREDIENT CALCULATION
# =============================================================================
def calculate_ingredients(weekly_fc: pd.DataFrame,
                           ingr: pd.DataFrame) -> pd.DataFrame:
    """
    For each SKU forecast, multiply predicted quantity × ingredient gram quantity.
    Returns aggregated ingredient requirements with safety buffer applied.
    """
    log.info("Calculating ingredient requirements …")

    merged = weekly_fc.merge(ingr[["pizza_name_id",
                                    "pizza_ingredients",
                                    "Items_Qty_In_Grams"]],
                              on="pizza_name_id",
                              how="left")

    missing_sku = merged[merged["pizza_ingredients"].isna()]["pizza_name_id"].unique()
    if len(missing_sku):
        log.warning("No ingredient data for %d SKU(s): %s",
                    len(missing_sku), list(missing_sku))

    merged = merged.dropna(subset=["pizza_ingredients"])
    merged["total_grams"] = (merged["predicted_qty_week"]
                             * merged["Items_Qty_In_Grams"])

    # Aggregate by ingredient
    agg = (merged.groupby("pizza_ingredients")["total_grams"]
                  .sum()
                  .reset_index()
                  .rename(columns={"pizza_ingredients": "ingredient",
                                    "total_grams"      : "required_grams"}))

    # Apply safety buffer
    agg["with_buffer_grams"] = (agg["required_grams"] * (1 + SAFETY_BUFFER)).round(0)
    agg["with_buffer_kg"]    = (agg["with_buffer_grams"] / 1000).round(3)
    agg = agg.sort_values("with_buffer_grams", ascending=False).reset_index(drop=True)

    log.info("Ingredient table: %d unique ingredients", len(agg))
    return agg, merged


# =============================================================================
# 7.  PURCHASE ORDER GENERATION
# =============================================================================
def generate_purchase_order(ingredients: pd.DataFrame,
                             weekly_fc: pd.DataFrame,
                             eval_df: pd.DataFrame,
                             daily: pd.DataFrame) -> None:
    """
    Write a formatted purchase order to Excel and CSV.
    Also saves weekly_forecast.csv and model_evaluation.csv.
    """
    log.info("Generating purchase order …")

    # ── Determine forecast period label ───────────────────────────────────────
    last_date   = daily["ds"].max()
    fc_start    = last_date + timedelta(days=1)
    fc_end      = fc_start  + timedelta(days=FORECAST_DAYS - 1)
    period_str  = f"{fc_start.strftime('%Y-%m-%d')}  to  {fc_end.strftime('%Y-%m-%d')}"

    # ── Evaluation CSV ────────────────────────────────────────────────────────
    eval_out = eval_df.copy()
    eval_out["mape_pct"] = (eval_out["mape"] * 100).round(2)
    eval_out.to_csv(OUTPUT_DIR / "model_evaluation.csv", index=False)
    log.info("  Saved: model_evaluation.csv")

    # ── Weekly forecast CSV ───────────────────────────────────────────────────
    weekly_fc.to_csv(OUTPUT_DIR / "weekly_forecast.csv", index=False)
    log.info("  Saved: weekly_forecast.csv")

    # ── Plain CSV purchase order ───────────────────────────────────────────────
    po_csv = ingredients[["ingredient",
                           "required_grams",
                           "with_buffer_grams",
                           "with_buffer_kg"]].copy()
    po_csv.insert(0, "rank", range(1, len(po_csv) + 1))
    po_csv.to_csv(OUTPUT_DIR / "purchase_order.csv", index=False)
    log.info("  Saved: purchase_order.csv")

    # ── Excel purchase order ───────────────────────────────────────────────────
    xlsx_path = OUTPUT_DIR / "purchase_order.xlsx"
    writer    = pd.ExcelWriter(xlsx_path, engine="openpyxl")

    # --- Sheet 1 : Cover / Summary -------------------------------------------
    valid_mapes = eval_df["mape"].dropna()
    summary_data = {
        "Field": [
            "Report Generated",
            "Forecast Period",
            "Forecast Horizon",
            "Total SKUs Modelled",
            "Prophet Models",
            "Mean-Fallback Models",
            "Overall MAPE (%)",
            "Safety Buffer Applied",
            "Total Ingredients",
        ],
        "Value": [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            period_str,
            f"{FORECAST_DAYS} days",
            len(eval_df),
            (eval_df["method"] == "prophet").sum(),
            (eval_df["method"] == "mean_fallback").sum(),
            f"{valid_mapes.mean() * 100:.2f}%" if len(valid_mapes) else "N/A",
            f"{SAFETY_BUFFER * 100:.0f}%",
            len(ingredients),
        ],
    }
    pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary",
                                         index=False)

    # --- Sheet 2 : Purchase Order ---------------------------------------------
    po_sheet = po_csv.rename(columns={
        "rank"              : "#",
        "ingredient"        : "Ingredient",
        "required_grams"    : "Base Requirement (g)",
        "with_buffer_grams" : f"Order Qty incl. {int(SAFETY_BUFFER*100)}% Buffer (g)",
        "with_buffer_kg"    : "Order Qty (kg)",
    })
    po_sheet.to_excel(writer, sheet_name="Purchase Order", index=False)

    # --- Sheet 3 : Weekly Forecast per SKU ------------------------------------
    weekly_fc_out = weekly_fc.copy()
    weekly_fc_out.columns = ["Pizza SKU", "Predicted Units (Next Week)"]
    weekly_fc_out.to_excel(writer, sheet_name="Forecast by SKU", index=False)

    # --- Sheet 4 : Model Evaluation -------------------------------------------
    eval_out_sheet = eval_out[["pizza_name_id","mape_pct","method"]].rename(columns={
        "pizza_name_id" : "Pizza SKU",
        "mape_pct"      : "MAPE (%)",
        "method"        : "Method",
    })
    eval_out_sheet.to_excel(writer, sheet_name="Model Evaluation", index=False)

    writer.close()

    # ── Format Excel with openpyxl ─────────────────────────────────────────────
    _format_excel(xlsx_path)
    log.info("  Saved: purchase_order.xlsx")


def _format_excel(xlsx_path: Path) -> None:
    """Apply professional formatting to the Excel workbook."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter

        wb = load_workbook(xlsx_path)

        header_fill  = PatternFill("solid", fgColor="1D3557")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        alt_fill     = PatternFill("solid", fgColor="EBF2F7")
        title_font   = Font(bold=True, size=13, color="1D3557")
        thin_border  = Border(
            bottom=Side(border_style="thin", color="CCCCCC")
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Header row
            for cell in ws[1]:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center",
                                            vertical="center", wrap_text=True)
                cell.border    = thin_border

            # Alternate row shading & alignment
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                fill = alt_fill if row_idx % 2 == 0 else PatternFill()
                for cell in row:
                    cell.fill      = fill
                    cell.alignment = Alignment(horizontal="left",
                                                vertical="center")
                    cell.border    = thin_border
                    # Format numbers
                    if isinstance(cell.value, float):
                        cell.number_format = "#,##0.000"
                    elif isinstance(cell.value, int):
                        cell.number_format = "#,##0"

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = (
                    min(max_len + 4, 50)
                )

            ws.row_dimensions[1].height = 28
            ws.freeze_panes = "A2"

        # Summary sheet title
        ws_sum = wb["Summary"]
        ws_sum["A1"].font = title_font

        wb.save(xlsx_path)
    except ImportError:
        log.warning("openpyxl formatting skipped (openpyxl not fully available).")
    except Exception as exc:
        log.warning("Excel formatting error (file still valid): %s", exc)


# =============================================================================
# 8.  INGREDIENT VISUALISATION
# =============================================================================
def plot_ingredients(ingredients: pd.DataFrame) -> None:
    """Bar chart of top ingredients by required quantity."""
    top = ingredients.head(20).copy()

    fig, ax = plt.subplots(figsize=(12, 7))
    bars = ax.barh(top["ingredient"], top["with_buffer_kg"],
                   color="#2A9D8F", edgecolor="white")
    ax.set(title=f"Top 20 Ingredients — Weekly Order (incl. {int(SAFETY_BUFFER*100)}% buffer)",
           xlabel="Quantity (kg)", ylabel="")
    ax.invert_yaxis()
    for bar, val in zip(bars, top["with_buffer_kg"]):
        ax.text(val + 0.05, bar.get_y() + bar.get_height()/2.,
                f"{val:.2f} kg", va="center", fontsize=8)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "eda_09_top_ingredients.png")
    plt.close()
    log.info("  Plot saved: eda_09_top_ingredients.png")


# =============================================================================
# MAIN
# =============================================================================
def main() -> None:
    log.info("=" * 60)
    log.info("  DOMINO'S PREDICTIVE PURCHASE ORDER SYSTEM")
    log.info("=" * 60)

    # 1. Load
    sales_raw, ingr_raw = load_data()

    # 2. Clean
    sales = clean_sales(sales_raw)
    ingr  = clean_ingredients(ingr_raw)

    # 3. EDA
    run_eda(sales)

    # 4. Feature engineering
    daily = build_daily_series(sales)

    # 5. Forecast
    weekly_fc, eval_df, forecast_full = train_and_forecast(daily)
    plot_forecast_sample(daily, forecast_full)

    # 6. Ingredients
    ingredients, _ = calculate_ingredients(weekly_fc, ingr)
    plot_ingredients(ingredients)

    # 7. Purchase order
    generate_purchase_order(ingredients, weekly_fc, eval_df, daily)

    log.info("")
    log.info("=" * 60)
    log.info("  ALL OUTPUTS SAVED TO: %s", OUTPUT_DIR.resolve())
    log.info("=" * 60)

    # ── Print summary to console ──────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  PURCHASE ORDER SUMMARY")
    print("=" * 60)
    print(f"\n  Forecast horizon   : {FORECAST_DAYS} days")
    valid = eval_df["mape"].dropna()
    if len(valid):
        print(f"  Overall MAPE       : {valid.mean()*100:.2f}%")
    print(f"  Total unique ingr. : {len(ingredients)}")
    print(f"\n  Top 10 ingredients to order:")
    for _, row in ingredients.head(10).iterrows():
        print(f"    {row['ingredient']:<35} {row['with_buffer_kg']:>8.3f} kg")
    print("\n  Files generated:")
    for f in sorted(OUTPUT_DIR.iterdir()):
        print(f"    {f.name}")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
